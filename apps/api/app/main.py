from __future__ import annotations

from enum import StrEnum
from datetime import date, datetime, timedelta, timezone
import csv
import hashlib
import io
import json
import os
import ipaddress
import re
import socket
import secrets
from pathlib import Path
from typing import Annotated, Literal
from uuid import UUID, uuid4

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, Response, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
import jwt
import httpx
from jwt import PyJWKClient
from pydantic import BaseModel, Field, HttpUrl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen.canvas import Canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display
from app.object_storage import put_object, scan_upload, validate_content_type
from app.persistence import EventStore, model_store, storage_health


class Role(StrEnum):
    PLATFORM_ADMIN = "platform_super_admin"
    ORG_ADMIN = "organization_admin"
    COMPLIANCE_MANAGER = "compliance_manager"
    ASSESSOR = "compliance_officer"
    AUDITOR = "external_auditor"
    VIEWER = "executive_viewer"


class UserContext(BaseModel):
    user_id: str
    organization_id: UUID
    role: Role


class AccountContext(BaseModel):
    user_id: UUID
    organization_id: UUID | None = None
    role: Role | None = None


class UserAccount(BaseModel):
    id: UUID
    email: str
    full_name: str
    password_hash: str
    mobile: str = ""
    job_title: str = ""
    preferred_language: Literal["ar", "en"] = "ar"
    active: bool = True
    created_at: datetime


class OrganizationMembership(BaseModel):
    id: UUID
    user_id: UUID
    organization_id: UUID
    role: Role = Role.ORG_ADMIN
    created_at: datetime


class ApplicationSession(BaseModel):
    id: UUID
    user_id: UUID
    token_hash: str
    expires_at: datetime
    created_at: datetime


class SignUpRequest(BaseModel):
    full_name: str = Field(min_length=2, max_length=160)
    email: str = Field(min_length=5, max_length=254)
    password: str = Field(min_length=10, max_length=128)


class SignInRequest(BaseModel):
    email: str
    password: str
    remember: bool = False


class ProfileUpdate(BaseModel):
    full_name: str = Field(min_length=2, max_length=160)
    mobile: str = Field(default="", max_length=40)
    job_title: str = Field(default="", max_length=120)
    preferred_language: Literal["ar", "en"] = "ar"


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str = Field(min_length=10, max_length=128)


class OrganizationProfile(BaseModel):
    entity_type: Literal["government", "semi_government", "private", "non_profit"]
    sector: str
    handles_personal_data: bool = False
    handles_sensitive_data: bool = False
    critical_infrastructure: bool = False
    critical_systems: bool = False
    uses_cloud: bool = False
    provides_cloud_services: bool = False
    sama_regulated: bool = False
    cma_regulated: bool = False
    subject_to_e_invoicing: bool = False
    seeks_iso_certification: bool = False
    provides_digital_services: bool = False
    has_cybersecurity_department: bool = False
    has_grc_team: bool = False
    has_branches: bool = False
    employee_size: str = ""
    existing_certifications: list[str] = Field(default_factory=list)
    current_status: dict[str, Literal["yes", "no", "partially", "unknown"]] = Field(default_factory=dict)
    objectives: list[str] = Field(default_factory=list)


class OrganizationCreate(BaseModel):
    name_ar: str = Field(min_length=2, max_length=200)
    name_en: str = Field(min_length=2, max_length=200)
    organization_type: str = "company"
    country: str = "Saudi Arabia"
    city: str = ""
    website: str = ""
    primary_contact: str = ""
    profile: OrganizationProfile


class Organization(OrganizationCreate):
    id: UUID
    onboarding_completed: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ApplicabilityResult(BaseModel):
    framework_code: str
    name_ar: str
    name_en: str
    classification: Literal["MANDATORY", "LIKELY_APPLICABLE", "CONDITIONAL", "VOLUNTARY", "NOT_APPLICABLE", "NEEDS_REVIEW"]
    reason_ar: str
    reason_en: str
    source_url: str


class AssessmentRequest(BaseModel):
    answers: dict[str, bool]
    framework_code: str | None = None


class EvidenceCreate(BaseModel):
    organization_id: UUID
    title: str = Field(min_length=2, max_length=250)
    universal_control_ids: list[str] = Field(min_length=1)
    classification: Literal["public", "internal", "confidential", "restricted"] = "internal"


class Evidence(EvidenceCreate):
    id: UUID
    state: Literal["uploaded", "under_review", "accepted", "rejected", "superseded", "archived"]
    sha256: str | None = None
    filename: str | None = None
    content_type: str | None = None
    size_bytes: int | None = None
    scan_status: Literal["not_applicable", "pending", "clean", "rejected", "failed", "unavailable"] = "not_applicable"
    effective_date: date | None = None
    expiry_date: date | None = None
    version: str = "1.0"
    replaces_evidence_id: UUID | None = None


class EvidenceLifecycleUpdate(BaseModel):
    state: Literal["under_review", "accepted", "rejected", "superseded", "archived"]
    expiry_date: date | None = None


class ActionCreate(BaseModel):
    organization_id: UUID
    gap_id: UUID | None = None
    title: str = Field(min_length=3, max_length=250)
    description: str = Field(default="", max_length=2000)
    owner: str = Field(min_length=2, max_length=120)
    due_date: date
    priority: Literal["critical", "high", "medium", "low"] = "medium"
    impacted_frameworks: list[str] = Field(default_factory=list)
    dependency: str | None = Field(default=None, max_length=500)
    completion_evidence_ids: list[UUID] = Field(default_factory=list)


class CorrectiveAction(ActionCreate):
    id: UUID
    status: Literal["open", "planned", "in_progress", "blocked", "pending_evidence", "pending_review", "completed", "accepted_risk"]


class ActionStatusUpdate(BaseModel):
    status: Literal["open", "planned", "in_progress", "blocked", "pending_evidence", "pending_review", "completed", "accepted_risk"]


class FrameworkMetric(BaseModel):
    code: str
    name_ar: str
    name_en: str
    score: int
    version: str


class DashboardSummary(BaseModel):
    organization_id: UUID
    overall_score: int
    evidence_readiness: int
    critical_gaps: int
    applicable_frameworks: int
    trend: float
    framework_scores: list[FrameworkMetric]
    actions: list[CorrectiveAction]
    risk_distribution: dict[str, int]
    disclaimer_ar: str


class EvidenceAnalysisRequest(BaseModel):
    requirement_reference: str = Field(min_length=2, max_length=100)


class EvidenceAnalysis(BaseModel):
    evidence_id: UUID
    suggestion: Literal["compliant", "partially_compliant", "insufficient", "missing"]
    confidence: float
    reasoning_ar: str
    citations: list[str]
    final_decision: None = None
    requires_human_approval: bool = True


class PolicyDraftRequest(BaseModel):
    organization_id: UUID
    policy_type: Literal["information_security", "access_control", "privacy", "data_retention", "business_continuity", "vendor_management", "ai_governance"]


class AuditEvent(BaseModel):
    organization_id: UUID
    actor_id: str
    action: str
    resource_type: str
    resource_id: str
    occurred_at: datetime


class WebsiteAuditRequest(BaseModel):
    url: HttpUrl


class JourneyReadinessRequest(BaseModel):
    completed_requirement_codes: list[str] = Field(default_factory=list, max_length=200)


class AssessmentCampaignCreate(BaseModel):
    organization_id: UUID
    framework_code: str = Field(min_length=2, max_length=80)
    framework_version: str = Field(min_length=1, max_length=40)
    title: str = Field(min_length=3, max_length=250)
    scope: str = Field(min_length=3, max_length=500)
    assessor_ids: list[str] = Field(default_factory=list)


class AssessmentCampaign(AssessmentCampaignCreate):
    id: UUID
    state: Literal["draft", "active", "under_review", "finalized"] = "draft"
    created_at: datetime


class AssessmentResponseUpsert(BaseModel):
    control_code: str = Field(min_length=2, max_length=100)
    status: Literal["compliant", "partially_compliant", "non_compliant", "not_assessed", "not_applicable"]
    score: float | None = Field(default=None, ge=0, le=100)
    rationale: str = Field(default="", max_length=2000)
    comment: str = Field(default="", max_length=2000)
    evidence_ids: list[UUID] = Field(default_factory=list)
    review_state: Literal["draft", "submitted", "approved", "changes_requested"] = "draft"
    mandatory: bool = False
    weight: float = Field(default=1, gt=0, le=100)


class AssessmentResponse(AssessmentResponseUpsert):
    id: UUID
    assessment_id: UUID
    organization_id: UUID
    assessor_id: str
    updated_at: datetime


class GapCreate(BaseModel):
    organization_id: UUID
    assessment_id: UUID | None = None
    framework_code: str
    control_code: str
    finding: str = Field(min_length=3, max_length=2000)
    current_state: str = Field(default="", max_length=2000)
    target_state: str = Field(default="", max_length=2000)
    severity: Literal["critical", "high", "medium", "low"]
    risk: str = Field(default="", max_length=1000)
    owner: str = Field(min_length=2, max_length=120)
    due_date: date | None = None
    remediation_plan: str = Field(default="", max_length=2000)


class Gap(GapCreate):
    id: UUID
    status: Literal["open", "in_progress", "blocked", "ready_for_review", "closed", "accepted_risk"] = "open"
    created_at: datetime


class GapStatusUpdate(BaseModel):
    status: Literal["open", "in_progress", "blocked", "ready_for_review", "closed", "accepted_risk"]


class ApplicabilityOverride(BaseModel):
    classification: Literal["MANDATORY", "LIKELY_APPLICABLE", "CONDITIONAL", "VOLUNTARY", "NOT_APPLICABLE", "NEEDS_REVIEW"]
    justification: str = Field(min_length=10, max_length=2000)


class ApplicabilityOverrideRecord(ApplicabilityOverride):
    id: UUID
    organization_id: UUID
    framework_code: str
    actor_id: str
    updated_at: datetime


class ControlMappingCreate(BaseModel):
    organization_id: UUID
    canonical_control_code: str = Field(min_length=2, max_length=100)
    framework_control_codes: list[str] = Field(min_length=1)
    mapping_type: Literal["equivalent", "partial", "supports", "related"]
    confidence: Literal["authoritative", "expert_reviewed", "ai_suggested", "unverified"]
    source: str = Field(min_length=3, max_length=500)
    version: str = Field(min_length=1, max_length=40)


class ControlMapping(ControlMappingCreate):
    id: UUID
    reviewer_id: str | None = None
    approved: bool = False


class MappingReview(BaseModel):
    decision: Literal["approved", "rejected", "under_review"]
    rationale: str = Field(min_length=5, max_length=1000)


class PolicyDocumentCreate(BaseModel):
    organization_id: UUID
    title_ar: str = Field(min_length=2, max_length=250)
    title_en: str = Field(min_length=2, max_length=250)
    document_type: Literal["policy", "procedure", "standard", "guideline", "template"]
    owner: str = Field(min_length=2, max_length=120)
    reviewer: str | None = None
    approver: str | None = None
    version: str = Field(min_length=1, max_length=40)
    effective_date: date | None = None
    next_review_date: date | None = None
    mapped_frameworks: list[str] = Field(default_factory=list)
    mapped_controls: list[str] = Field(default_factory=list)
    attachment_ids: list[UUID] = Field(default_factory=list)
    ai_assisted: bool = False


class PolicyDocument(PolicyDocumentCreate):
    id: UUID
    status: Literal["draft", "under_review", "approved", "published", "superseded", "archived"] = "draft"
    created_at: datetime
    updated_at: datetime


class PolicyTransition(BaseModel):
    status: Literal["under_review", "approved", "published", "superseded", "archived", "draft"]
    comment: str = Field(default="", max_length=1000)


class Notification(BaseModel):
    id: UUID
    organization_id: UUID
    recipient_id: str
    notification_type: str
    title_ar: str
    title_en: str
    resource_type: str
    resource_id: str
    resource_url: str
    severity: Literal["info", "medium", "high", "critical"] = "info"
    read_at: datetime | None = None
    created_at: datetime


class ComplianceSnapshot(BaseModel):
    id: UUID
    organization_id: UUID
    overall_readiness: float
    framework_readiness: dict[str, float] = Field(default_factory=dict)
    domain_readiness: dict[str, float] = Field(default_factory=dict)
    open_critical_gaps: int
    overdue_actions: int
    evidence_coverage: float
    reason: str
    captured_at: datetime


class KnowledgeSourceCreate(BaseModel):
    organization_id: UUID | None = None
    title: str = Field(min_length=3, max_length=250)
    source_url: HttpUrl
    framework_code: str
    framework_version: str
    source_status: Literal["official", "verified", "under_review", "demo_unverified", "superseded"]
    content: str = Field(min_length=20, max_length=100000)


class KnowledgeSource(KnowledgeSourceCreate):
    id: UUID
    checksum: str
    chunks: list[str]
    ingested_at: datetime


class AssistantQuery(BaseModel):
    question: str = Field(min_length=3, max_length=1000)
    framework_code: str | None = None


APP_ENV = os.getenv("APP_ENV", "development").lower()
IS_PRODUCTION = APP_ENV == "production"
ALLOWED_ORIGINS = [origin.strip() for origin in os.getenv("ALLOWED_ORIGINS", "http://localhost:3000").split(",") if origin.strip()]
OIDC_ISSUER = os.getenv("OIDC_ISSUER", "").rstrip("/")
OIDC_AUDIENCE = os.getenv("OIDC_AUDIENCE", "")
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "/tmp/multazim-evidence" if IS_PRODUCTION else ".data/evidence"))
ALLOWED_EVIDENCE_TYPES = {"application/pdf", "image/png", "image/jpeg", "text/csv", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}


def arabic_pdf_font() -> str:
    candidates = [os.getenv("ARABIC_FONT_PATH", ""), "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arabtype.ttf"]
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            if "MultazimArabic" not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont("MultazimArabic", candidate))
            return "MultazimArabic"
    raise RuntimeError("No Arabic-capable PDF font found; configure ARABIC_FONT_PATH")


def rtl(text: str) -> str:
    return get_display(arabic_reshaper.reshape(text))

app = FastAPI(
    title="Multazim Compliance Intelligence API",
    version="0.2.0",
    description="Tenant-aware compliance API with OIDC enforcement in production and explicit demo mode in development.",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-User-Id", "X-Organization-Id", "X-Role"],
)

organizations = model_store("organizations", Organization)
user_store = model_store("user_accounts", UserAccount)
membership_store = model_store("organization_memberships", OrganizationMembership)
session_store = model_store("application_sessions", ApplicationSession)
evidence_store = model_store("evidence", Evidence)
DEMO_ORGANIZATION_ID = UUID("11111111-1111-4111-8111-111111111111")
organizations[DEMO_ORGANIZATION_ID] = Organization(
    id=DEMO_ORGANIZATION_ID,
    name_ar="شركة آفاق الرقمية السعودية",
    name_en="Saudi Digital Horizons Company",
    organization_type="company",
    profile=OrganizationProfile(entity_type="government", sector="technology", handles_personal_data=True, uses_cloud=True),
    onboarding_completed=True,
)
action_store = model_store("actions", CorrectiveAction)
audit_events = EventStore(AuditEvent)
assessment_store = model_store("assessment_campaigns", AssessmentCampaign)
assessment_response_store = model_store("assessment_responses", AssessmentResponse)
gap_store = model_store("gaps", Gap)
mapping_store = model_store("control_mappings", ControlMapping)
applicability_override_store = model_store("applicability_overrides", ApplicabilityOverrideRecord)
policy_store = model_store("policy_documents", PolicyDocument)
notification_store = model_store("notifications", Notification)
snapshot_store = model_store("compliance_snapshots", ComplianceSnapshot)
knowledge_store = model_store("knowledge_sources", KnowledgeSource)
for item in [
    CorrectiveAction(id=UUID("21111111-1111-4111-8111-111111111111"), organization_id=DEMO_ORGANIZATION_ID, title="اعتماد مراجعة الحسابات ذات الصلاحيات العالية", owner="نورة القحطاني", due_date=date(2026, 8, 12), priority="critical", impacted_frameworks=["NCA ECC", "ISO 27001", "SAMA CSF", "CST CRF"], status="open"),
    CorrectiveAction(id=UUID("21111111-1111-4111-8111-111111111112"), organization_id=DEMO_ORGANIZATION_ID, title="استكمال سجل أنشطة معالجة البيانات", owner="فريق الخصوصية", due_date=date(2026, 8, 17), priority="high", impacted_frameworks=["PDPL", "ISO 27701"], status="in_progress"),
    CorrectiveAction(id=UUID("21111111-1111-4111-8111-111111111113"), organization_id=DEMO_ORGANIZATION_ID, title="توثيق اختبار استعادة النسخ الاحتياطية", owner="إدارة التقنية", due_date=date(2026, 8, 25), priority="medium", impacted_frameworks=["DGA Qiyas", "ISO 22301", "SAMA BCM"], status="planned"),
]:
    action_store[item.id] = item


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    response.headers["Content-Security-Policy"] = "default-src 'self'; frame-ancestors 'none'; object-src 'none'; base-uri 'self'"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains" if IS_PRODUCTION else "max-age=0"
    return response


def normalize_email(value: str) -> str:
    email = value.strip().lower()
    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
        raise HTTPException(status_code=422, detail="Enter a valid work email")
    return email


def hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 600_000)
    return f"pbkdf2_sha256$600000${salt.hex()}${digest.hex()}"


def verify_password(password: str, encoded: str) -> bool:
    try:
        algorithm, iterations, salt, expected = encoded.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        actual = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), int(iterations)).hex()
        return secrets.compare_digest(actual, expected)
    except (ValueError, TypeError):
        return False


def create_session(user_id: UUID, remember: bool = False) -> tuple[str, ApplicationSession]:
    token = f"mz_{secrets.token_urlsafe(48)}"
    now = datetime.now(timezone.utc)
    item = ApplicationSession(id=uuid4(), user_id=user_id, token_hash=hashlib.sha256(token.encode()).hexdigest(),
        expires_at=now + timedelta(days=30 if remember else 1), created_at=now)
    session_store[item.id] = item
    return token, item


def account_context(authorization: Annotated[str | None, Header()] = None) -> AccountContext:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Authentication required")
    token = authorization.split(" ", 1)[1]
    if not token.startswith("mz_"):
        raise HTTPException(status_code=401, detail="Invalid application session")
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    session = next((item for item in session_store.values() if secrets.compare_digest(item.token_hash, token_hash)), None)
    if not session or session.expires_at <= datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    user = user_store.get(session.user_id)
    if not user or not user.active:
        raise HTTPException(status_code=401, detail="Account is unavailable")
    membership = next((item for item in membership_store.values() if item.user_id == user.id), None)
    return AccountContext(user_id=user.id, organization_id=membership.organization_id if membership else None,
        role=membership.role if membership else None)


def user_context(
    authorization: Annotated[str | None, Header()] = None,
    x_user_id: Annotated[str | None, Header()] = None,
    x_organization_id: Annotated[UUID | None, Header()] = None,
    x_role: Annotated[Role | None, Header()] = None,
) -> UserContext:
    if authorization and authorization.lower().startswith("bearer "):
        token = authorization.split(" ", 1)[1]
        if token.startswith("mz_"):
            account = account_context(authorization)
            if not account.organization_id or not account.role:
                raise HTTPException(status_code=409, detail="Create an organization before using this module")
            return UserContext(user_id=str(account.user_id), organization_id=account.organization_id, role=account.role)
        if not OIDC_ISSUER or not OIDC_AUDIENCE:
            raise HTTPException(status_code=503, detail="OIDC is not configured")
        try:
            signing_key = PyJWKClient(f"{OIDC_ISSUER}/.well-known/jwks.json").get_signing_key_from_jwt(token)
            claims = jwt.decode(token, signing_key.key, algorithms=["RS256"], audience=OIDC_AUDIENCE, issuer=OIDC_ISSUER)
            return UserContext(user_id=str(claims["sub"]), organization_id=UUID(str(claims["organization_id"])), role=Role(claims["role"]))
        except (KeyError, ValueError, jwt.PyJWTError) as exc:
            raise HTTPException(status_code=401, detail="Invalid access token") from exc
    if not IS_PRODUCTION and x_user_id and x_organization_id and x_role:
        return UserContext(user_id=x_user_id, organization_id=x_organization_id, role=x_role)
    raise HTTPException(status_code=401, detail="Bearer authentication required")


def require_roles(*roles: Role):
    def checker(user: Annotated[UserContext, Depends(user_context)]) -> UserContext:
        if user.role not in roles:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient role")
        return user
    return checker


def assert_tenant(resource_org_id: UUID, user: UserContext) -> None:
    if user.role != Role.PLATFORM_ADMIN and resource_org_id != user.organization_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Resource not found")


def record_event(user: UserContext, action: str, resource_type: str, resource_id: UUID) -> None:
    audit_events.append(AuditEvent(organization_id=user.organization_id, actor_id=user.user_id, action=action,
        resource_type=resource_type, resource_id=str(resource_id), occurred_at=datetime.now(timezone.utc)))


def notify(user: UserContext, notification_type: str, title_ar: str, title_en: str,
    resource_type: str, resource_id: UUID, resource_url: str, severity: str = "info") -> Notification:
    existing = next((item for item in notification_store.values() if item.organization_id == user.organization_id
        and item.recipient_id == user.user_id and item.notification_type == notification_type
        and item.resource_type == resource_type and item.resource_id == str(resource_id)), None)
    item = Notification(id=existing.id if existing else uuid4(), organization_id=user.organization_id,
        recipient_id=user.user_id, notification_type=notification_type, title_ar=title_ar, title_en=title_en,
        resource_type=resource_type, resource_id=str(resource_id), resource_url=resource_url,
        severity=severity, read_at=None, created_at=datetime.now(timezone.utc))
    notification_store[item.id] = item
    return item


def capture_snapshot(user: UserContext, reason: str, readiness: float | None = None) -> ComplianceSnapshot:
    tenant_gaps = [item for item in gap_store.values() if item.organization_id == user.organization_id]
    tenant_actions = [item for item in action_store.values() if item.organization_id == user.organization_id]
    tenant_evidence = [item for item in evidence_store.values() if item.organization_id == user.organization_id and item.state != "archived"]
    accepted = sum(item.state == "accepted" for item in tenant_evidence)
    item = ComplianceSnapshot(id=uuid4(), organization_id=user.organization_id,
        overall_readiness=float(readiness if readiness is not None else 0), framework_readiness={}, domain_readiness={},
        open_critical_gaps=sum(gap.severity == "critical" and gap.status not in {"closed", "accepted_risk"} for gap in tenant_gaps),
        overdue_actions=sum(action.due_date < date.today() and action.status != "completed" for action in tenant_actions),
        evidence_coverage=round(accepted / len(tenant_evidence) * 100, 1) if tenant_evidence else 0,
        reason=reason, captured_at=datetime.now(timezone.utc))
    snapshot_store[item.id] = item
    return item


def determine_applicability(profile: OrganizationProfile) -> list[ApplicabilityResult]:
    results: list[ApplicabilityResult] = []

    def add(code: str, ar: str, en: str, level: str, why_ar: str, why_en: str, source: str):
        results.append(ApplicabilityResult(framework_code=code, name_ar=ar, name_en=en,
            classification=level, reason_ar=why_ar, reason_en=why_en, source_url=source))

    if profile.entity_type == "government":
        add("DGA-QIYAS-2025", "المعايير الأساسية للتحول الرقمي 2025", "Digital Transformation Basic Standards 2025", "MANDATORY",
            "الجهة مصنفة كجهة حكومية ضمن نطاق قياس.", "The organization is classified as a government entity within Qiyas scope.", "https://dga.gov.sa/en/Standards_Of_Digital_Transformation")
    else:
        add("DGA-QIYAS-2025", "المعايير الأساسية للتحول الرقمي 2025", "Digital Transformation Basic Standards 2025", "NOT_APPLICABLE",
            "قياس موجه للجهات الحكومية؛ يلزم التحقق إذا كان للجهة تكليف خاص.", "Qiyas targets government entities; review any specific mandate separately.", "https://dga.gov.sa/en/Standards_Of_Digital_Transformation")
    if profile.handles_personal_data:
        add("SDAIA-PDPL", "نظام حماية البيانات الشخصية", "Personal Data Protection Law", "MANDATORY",
            "أفادت الجهة بأنها تعالج بيانات شخصية.", "The organization declared that it processes personal data.", "https://sdaia.gov.sa/en/SDAIA/about/Pages/RegulationsAndPolicies.aspx")
    if profile.sama_regulated:
        add("SAMA-CSF", "إطار الأمن السيبراني", "SAMA Cyber Security Framework", "MANDATORY",
            "أفادت الجهة بأنها خاضعة لرقابة البنك المركزي السعودي.", "The organization declared that it is regulated by the Saudi Central Bank.", "https://rulebook.sama.gov.sa/en/cyber-security-framework-2")
    if profile.uses_cloud:
        add("NCA-CCC-2-2024", "ضوابط الأمن السيبراني للحوسبة السحابية", "Cloud Cybersecurity Controls", "CONDITIONAL",
            "تستخدم الجهة خدمات سحابية؛ يجب تحديد دور المستفيد أو مقدم الخدمة.", "The organization uses cloud services; tenant/provider role must be confirmed.", "https://nca.gov.sa/en/regulatory-documents/controls-list/ccc/")
    add("NCA-ECC-2-2024", "الضوابط الأساسية للأمن السيبراني", "Essential Cybersecurity Controls", "NEEDS_REVIEW",
        "يتطلب تحديد نطاق انطباق الضوابط حسب طبيعة الجهة وأصولها.", "Applicability requires review of the entity mandate and information assets.", "https://nca.gov.sa/en/regulatory-documents/controls-list/ecc/")
    add("ISO-27001-2022", "نظام إدارة أمن المعلومات", "ISO/IEC 27001", "VOLUNTARY" if profile.seeks_iso_certification else "CONDITIONAL",
        "معيار شهادة دولي وليس متطلبًا سعوديًا افتراضيًا.", "An international certification standard, not a Saudi requirement by default.", "https://www.iso.org/standard/27001")
    return results


@app.get("/health")
def health():
    oidc_ready = bool(OIDC_ISSUER and OIDC_AUDIENCE)
    return {"status": "ok", "service": "multazim-api", "version": app.version, "environment": APP_ENV,
        "api_url": os.getenv("PUBLIC_API_URL", "http://localhost:8000"), "storage": storage_health(),
        "checks": {"database_configured": True, "oidc_configured": oidc_ready,
        "persistent_object_storage": True, "demo_headers_enabled": not IS_PRODUCTION}}


def account_payload(account: AccountContext) -> dict[str, object]:
    user = user_store[account.user_id]
    organization = organizations.get(account.organization_id) if account.organization_id else None
    membership = next((item for item in membership_store.values() if item.user_id == user.id), None)
    return {"user": {"id": str(user.id), "email": user.email, "full_name": user.full_name, "mobile": user.mobile,
        "job_title": user.job_title, "preferred_language": user.preferred_language, "created_at": user.created_at},
        "organization": organization.model_dump(mode="json") if organization else None,
        "membership": membership.model_dump(mode="json") if membership else None}


@app.post("/v1/auth/signup", status_code=201)
def sign_up(payload: SignUpRequest):
    email = normalize_email(payload.email)
    if any(item.email == email for item in user_store.values()):
        raise HTTPException(status_code=409, detail="An account with this email already exists")
    user = UserAccount(id=uuid4(), email=email, full_name=payload.full_name.strip(),
        password_hash=hash_password(payload.password), created_at=datetime.now(timezone.utc))
    user_store[user.id] = user
    token, session = create_session(user.id)
    return {"token": token, "expires_at": session.expires_at, **account_payload(AccountContext(user_id=user.id))}


@app.post("/v1/auth/signin")
def sign_in(payload: SignInRequest):
    email = normalize_email(payload.email)
    user = next((item for item in user_store.values() if item.email == email), None)
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Email or password is incorrect")
    token, session = create_session(user.id, payload.remember)
    membership = next((item for item in membership_store.values() if item.user_id == user.id), None)
    return {"token": token, "expires_at": session.expires_at,
        **account_payload(AccountContext(user_id=user.id, organization_id=membership.organization_id if membership else None,
            role=membership.role if membership else None))}


@app.post("/v1/auth/logout", status_code=204)
def logout(authorization: Annotated[str | None, Header()] = None):
    if authorization and authorization.lower().startswith("bearer "):
        token_hash = hashlib.sha256(authorization.split(" ", 1)[1].encode()).hexdigest()
        session = next((item for item in session_store.values() if secrets.compare_digest(item.token_hash, token_hash)), None)
        if session:
            del session_store[session.id]
    return Response(status_code=204)


@app.get("/v1/me")
def current_account(account: Annotated[AccountContext, Depends(account_context)]):
    return account_payload(account)


@app.put("/v1/me/profile")
def update_profile(payload: ProfileUpdate, account: Annotated[AccountContext, Depends(account_context)]):
    user = user_store[account.user_id]
    user_store[user.id] = user.model_copy(update=payload.model_dump())
    return account_payload(account)


@app.put("/v1/me/password", status_code=204)
def change_password(payload: ChangePasswordRequest, account: Annotated[AccountContext, Depends(account_context)]):
    user = user_store[account.user_id]
    if not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    user_store[user.id] = user.model_copy(update={"password_hash": hash_password(payload.new_password)})
    return Response(status_code=204)


@app.post("/v1/onboarding/organization", response_model=Organization, status_code=201)
def create_onboarding_organization(payload: OrganizationCreate,
    account: Annotated[AccountContext, Depends(account_context)]):
    existing = next((item for item in membership_store.values() if item.user_id == account.user_id), None)
    if existing:
        raise HTTPException(status_code=409, detail="This account already belongs to an organization")
    organization = Organization(id=uuid4(), onboarding_completed=True, **payload.model_dump())
    organizations[organization.id] = organization
    membership = OrganizationMembership(id=uuid4(), user_id=account.user_id, organization_id=organization.id,
        role=Role.ORG_ADMIN, created_at=datetime.now(timezone.utc))
    membership_store[membership.id] = membership
    return organization


@app.get("/v1/organization/workspace")
def organization_workspace(user: Annotated[UserContext, Depends(user_context)]):
    organization = organizations.get(user.organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organization not found")
    applicability = determine_applicability(organization.profile)
    tenant_assessments = [item for item in assessment_store.values() if item.organization_id == user.organization_id]
    tenant_actions = [item for item in action_store.values() if item.organization_id == user.organization_id and item.status != "completed"]
    recent = sorted((item for item in audit_events if item.organization_id == user.organization_id), key=lambda item: item.occurred_at, reverse=True)[:5]
    completed = 5 if organization.onboarding_completed else 1
    if tenant_assessments:
        completed += 1
    if any(item.organization_id == user.organization_id for item in evidence_store.values()):
        completed += 1
    return {"organization": organization, "onboarding_completeness": 100 if organization.onboarding_completed else 25,
        "readiness": next((item.overall_readiness for item in sorted(snapshot_store.values(), key=lambda x: x.captured_at, reverse=True)
            if item.organization_id == user.organization_id), 0), "active_journeys": len(tenant_assessments),
        "pending_actions": len(tenant_actions), "getting_started_completed": completed,
        "recommendations": [item.model_dump(mode="json") for item in applicability if item.classification != "NOT_APPLICABLE"],
        "recent_activity": [item.model_dump(mode="json") for item in recent]}


@app.post("/v1/organizations", response_model=Organization, status_code=201)
def create_organization(payload: OrganizationCreate):
    organization = Organization(id=uuid4(), **payload.model_dump())
    organizations[organization.id] = organization
    return organization


@app.put("/v1/organizations/{organization_id}/profile", response_model=Organization)
def update_organization_profile(
    organization_id: UUID,
    profile: OrganizationProfile,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))],
):
    assert_tenant(organization_id, user)
    organization = organizations.get(organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Resource not found")
    updated = organization.model_copy(update={"profile": profile})
    organizations[organization_id] = updated
    record_event(user, "profile.updated", "organization", organization_id)
    return updated


@app.post("/v1/applicability", response_model=list[ApplicabilityResult])
def applicability(profile: OrganizationProfile):
    return determine_applicability(profile)


@app.get("/v1/organizations/{organization_id}/applicability", response_model=list[ApplicabilityResult])
def organization_applicability(organization_id: UUID, user: Annotated[UserContext, Depends(user_context)]):
    assert_tenant(organization_id, user)
    organization = organizations.get(organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Resource not found")
    return determine_applicability(organization.profile)


@app.put("/v1/organizations/{organization_id}/applicability/{framework_code}/override", response_model=ApplicabilityOverrideRecord)
def override_applicability(organization_id: UUID, framework_code: str, payload: ApplicabilityOverride,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))]):
    assert_tenant(organization_id, user)
    existing = next((item for item in applicability_override_store.values()
        if item.organization_id == organization_id and item.framework_code == framework_code), None)
    record = ApplicabilityOverrideRecord(id=existing.id if existing else uuid4(), organization_id=organization_id,
        framework_code=framework_code, actor_id=user.user_id, updated_at=datetime.now(timezone.utc), **payload.model_dump())
    applicability_override_store[record.id] = record
    record_event(user, "applicability.overridden", "framework", record.id)
    return record


def tenant_assessment(assessment_id: UUID, user: UserContext) -> AssessmentCampaign:
    item = assessment_store.get(assessment_id)
    if not item:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(item.organization_id, user)
    return item


@app.get("/v1/assessments", response_model=list[AssessmentCampaign])
def list_assessments(user: Annotated[UserContext, Depends(user_context)]):
    return [item for item in assessment_store.values() if item.organization_id == user.organization_id]


@app.post("/v1/assessments", response_model=AssessmentCampaign, status_code=201)
def create_assessment(payload: AssessmentCampaignCreate,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))]):
    assert_tenant(payload.organization_id, user)
    item = AssessmentCampaign(id=uuid4(), created_at=datetime.now(timezone.utc), **payload.model_dump())
    assessment_store[item.id] = item
    record_event(user, "assessment.created", "assessment", item.id)
    notify(user, "assessment_assigned", "تم إسناد تقييم جديد", "New assessment assigned", "assessment", item.id, "/assessment", "medium")
    return item


@app.get("/v1/assessments/{assessment_id}/responses", response_model=list[AssessmentResponse])
def list_assessment_responses(assessment_id: UUID, user: Annotated[UserContext, Depends(user_context)]):
    tenant_assessment(assessment_id, user)
    return [item for item in assessment_response_store.values() if item.assessment_id == assessment_id]


@app.put("/v1/assessments/{assessment_id}/responses/{control_code}", response_model=AssessmentResponse)
def upsert_assessment_response(assessment_id: UUID, control_code: str, payload: AssessmentResponseUpsert,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR))]):
    campaign = tenant_assessment(assessment_id, user)
    if payload.control_code != control_code:
        raise HTTPException(status_code=422, detail="Control code mismatch")
    for evidence_id in payload.evidence_ids:
        evidence = evidence_store.get(evidence_id)
        if not evidence or evidence.organization_id != user.organization_id or evidence.state == "rejected":
            raise HTTPException(status_code=422, detail="Evidence is unavailable for this organization")
    existing = next((item for item in assessment_response_store.values()
        if item.assessment_id == assessment_id and item.control_code == control_code), None)
    response = AssessmentResponse(id=existing.id if existing else uuid4(), assessment_id=assessment_id,
        organization_id=campaign.organization_id, assessor_id=user.user_id, updated_at=datetime.now(timezone.utc), **payload.model_dump())
    assessment_response_store[response.id] = response
    record_event(user, "assessment.response_updated", "assessment_response", response.id)
    return response


@app.get("/v1/assessments/{assessment_id}/score")
def assessment_score(assessment_id: UUID, user: Annotated[UserContext, Depends(user_context)]):
    tenant_assessment(assessment_id, user)
    responses = [item for item in assessment_response_store.values() if item.assessment_id == assessment_id]
    assessed = [item for item in responses if item.status not in {"not_assessed", "not_applicable"}]
    applicable = [item for item in responses if item.status != "not_applicable"]
    denominator = sum(item.weight for item in assessed)
    earned = sum(item.weight * ({"compliant": 1, "partially_compliant": .5, "non_compliant": 0}.get(item.status, 0)) for item in assessed)
    mandatory_failures = sum(item.mandatory and item.status == "non_compliant" for item in assessed)
    raw = round(earned / denominator * 100, 1) if denominator else None
    penalty = min(mandatory_failures * 5, 25)
    readiness = max(0, raw - penalty) if raw is not None else None
    completeness = round(len(assessed) / len(applicable) * 100, 1) if applicable else 0
    capture_snapshot(user, "assessment_recalculated", readiness)
    return {"readiness": readiness, "raw_score": raw, "mandatory_control_penalty": penalty,
        "assessment_completeness": completeness, "assessed_controls": len(assessed),
        "applicable_controls": len(applicable), "not_applicable": len(responses) - len(applicable),
        "methodology": "weighted-status-v1", "explanation": "Not assessed controls affect completeness, not readiness; N/A controls are excluded; mandatory failures subtract 5 points each (maximum 25)."}


@app.get("/v1/dashboard", response_model=DashboardSummary)
def dashboard(user: Annotated[UserContext, Depends(user_context)]):
    tenant_actions = [item for item in action_store.values() if item.organization_id == user.organization_id]
    tenant_assessments = [item for item in assessment_store.values() if item.organization_id == user.organization_id]
    tenant_responses = [item for item in assessment_response_store.values() if item.organization_id == user.organization_id]
    tenant_evidence = [item for item in evidence_store.values() if item.organization_id == user.organization_id and item.state not in {"archived", "superseded"}]
    tenant_gaps = [item for item in gap_store.values() if item.organization_id == user.organization_id and item.status not in {"closed", "accepted_risk"}]
    framework_names = {
        "DGA-QIYAS-2025": ("معايير التحول الرقمي الأساسية", "Digital Transformation Basic Standards"),
        "NCA-ECC-2-2024": ("الضوابط الأساسية للأمن السيبراني", "Essential Cybersecurity Controls"),
        "SDAIA-PDPL": ("نظام حماية البيانات الشخصية", "Personal Data Protection Law"),
        "ISO-27001-2022": ("نظام إدارة أمن المعلومات", "ISO/IEC 27001"),
    }
    weights = {"compliant": 100, "partially_compliant": 50, "non_compliant": 0}
    framework_scores = []
    for assessment in tenant_assessments:
        responses = [item for item in tenant_responses if item.assessment_id == assessment.id and item.status in weights]
        if not responses:
            continue
        total_weight = sum(item.weight for item in responses)
        score = round(sum(weights[item.status] * item.weight for item in responses) / total_weight) if total_weight else 0
        names = framework_names.get(assessment.framework_code, (assessment.framework_code, assessment.framework_code))
        framework_scores.append(FrameworkMetric(code=assessment.framework_code, name_ar=names[0], name_en=names[1],
            score=score, version=assessment.framework_version))
    history = sorted((item for item in snapshot_store.values() if item.organization_id == user.organization_id), key=lambda item: item.captured_at)
    overall = round(sum(item.score for item in framework_scores) / len(framework_scores)) if framework_scores else 0
    evidence_readiness = round(sum(item.state == "accepted" for item in tenant_evidence) / len(tenant_evidence) * 100) if tenant_evidence else 0
    organization = organizations.get(user.organization_id)
    applicable = len([item for item in determine_applicability(organization.profile) if item.classification != "NOT_APPLICABLE"]) if organization else 0
    return DashboardSummary(
        organization_id=user.organization_id, overall_score=overall, evidence_readiness=evidence_readiness,
        critical_gaps=sum(item.priority == "critical" and item.status != "completed" for item in tenant_actions),
        applicable_frameworks=applicable, trend=(history[-1].overall_readiness - history[-2].overall_readiness) if len(history) >= 2 else 0,
        framework_scores=framework_scores, actions=tenant_actions,
        risk_distribution={level: sum(item.severity == level for item in tenant_gaps) for level in ("critical", "high", "medium")},
        disclaimer_ar="درجات ملتزم مؤشرات داخلية مبنية على استجابات المؤسسة، وليست تقييمًا رسميًا صادرًا من جهة تنظيمية.",
    )


@app.get("/v1/actions", response_model=list[CorrectiveAction])
def list_actions(user: Annotated[UserContext, Depends(user_context)]):
    return [item for item in action_store.values() if item.organization_id == user.organization_id]


@app.get("/v1/gaps", response_model=list[Gap])
def list_gaps(user: Annotated[UserContext, Depends(user_context)], overdue: bool = False):
    items = [item for item in gap_store.values() if item.organization_id == user.organization_id]
    if overdue:
        items = [item for item in items if item.due_date and item.due_date < date.today() and item.status not in {"closed", "accepted_risk"}]
    return items


@app.post("/v1/gaps", response_model=Gap, status_code=201)
def create_gap(payload: GapCreate,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR))]):
    assert_tenant(payload.organization_id, user)
    if payload.assessment_id:
        tenant_assessment(payload.assessment_id, user)
    item = Gap(id=uuid4(), created_at=datetime.now(timezone.utc), **payload.model_dump())
    gap_store[item.id] = item
    record_event(user, "gap.created", "gap", item.id)
    notify(user, "gap_assigned", "تم إسناد فجوة امتثال", "Compliance gap assigned", "gap", item.id, "/gaps", "info" if item.severity == "low" else item.severity)
    return item


@app.patch("/v1/gaps/{gap_id}", response_model=Gap)
def update_gap(gap_id: UUID, payload: GapStatusUpdate,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR))]):
    item = gap_store.get(gap_id)
    if not item:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(item.organization_id, user)
    updated = item.model_copy(update={"status": payload.status})
    gap_store[gap_id] = updated
    record_event(user, "gap.status_updated", "gap", gap_id)
    notify(user, "gap_status_changed", "تغيرت حالة فجوة", "Gap status changed", "gap", gap_id, "/gaps", "info" if item.severity == "low" else item.severity)
    if payload.status == "closed":
        capture_snapshot(user, "gap_closed")
    return updated


@app.get("/v1/control-mappings", response_model=list[ControlMapping])
def list_control_mappings(user: Annotated[UserContext, Depends(user_context)]):
    return [item for item in mapping_store.values() if item.organization_id == user.organization_id]


@app.post("/v1/control-mappings", response_model=ControlMapping, status_code=201)
def create_control_mapping(payload: ControlMappingCreate,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))]):
    assert_tenant(payload.organization_id, user)
    approved = payload.confidence in {"authoritative", "expert_reviewed"}
    item = ControlMapping(id=uuid4(), reviewer_id=user.user_id if approved else None, approved=approved, **payload.model_dump())
    mapping_store[item.id] = item
    record_event(user, "control_mapping.created", "control_mapping", item.id)
    return item


@app.patch("/v1/control-mappings/{mapping_id}/review", response_model=ControlMapping)
def review_control_mapping(mapping_id: UUID, payload: MappingReview,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))]):
    item = mapping_store.get(mapping_id)
    if not item:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(item.organization_id, user)
    approved = payload.decision == "approved"
    if approved and item.confidence in {"ai_suggested", "unverified"} and not payload.rationale:
        raise HTTPException(status_code=422, detail="Human review rationale is required")
    updated = item.model_copy(update={"approved": approved, "reviewer_id": user.user_id})
    mapping_store[mapping_id] = updated
    record_event(user, f"control_mapping.{payload.decision}", "control_mapping", mapping_id)
    return updated


@app.get("/v1/evidence/{evidence_id}/coverage")
def evidence_coverage(evidence_id: UUID, user: Annotated[UserContext, Depends(user_context)]):
    evidence = evidence_store.get(evidence_id)
    if not evidence:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(evidence.organization_id, user)
    mappings = [item for item in mapping_store.values() if item.organization_id == user.organization_id
        and item.approved and item.canonical_control_code in evidence.universal_control_ids]
    controls = sorted({code for item in mappings for code in item.framework_control_codes})
    frameworks = sorted({code.split("-")[0] for code in controls})
    return {"evidence_id": evidence_id, "supported_requirements": controls,
        "requirements_count": len(controls), "frameworks": frameworks, "frameworks_count": len(frameworks),
        "authoritative": all(item.confidence == "authoritative" for item in mappings) if mappings else False}


@app.post("/v1/actions", response_model=CorrectiveAction, status_code=201)
def create_action(payload: ActionCreate, user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR))]):
    assert_tenant(payload.organization_id, user)
    if payload.gap_id:
        gap = gap_store.get(payload.gap_id)
        if not gap:
            raise HTTPException(status_code=404, detail="Resource not found")
        assert_tenant(gap.organization_id, user)
    item = CorrectiveAction(id=uuid4(), status="open", **payload.model_dump())
    action_store[item.id] = item
    record_event(user, "action.created", "corrective_action", item.id)
    notify(user, "action_due", "تم إسناد إجراء تصحيحي", "Corrective action assigned", "action", item.id, "/gaps", "high" if item.priority in {"critical", "high"} else "medium")
    return item


@app.patch("/v1/actions/{action_id}", response_model=CorrectiveAction)
def update_action(action_id: UUID, payload: ActionStatusUpdate, user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR))]):
    item = action_store.get(action_id)
    if not item:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(item.organization_id, user)
    updated = item.model_copy(update={"status": payload.status})
    action_store[action_id] = updated
    record_event(user, "action.status_updated", "corrective_action", action_id)
    return updated


@app.post("/v1/assessments/score")
def score_assessment(payload: AssessmentRequest):
    weights_by_framework = {
        "NCA-ECC-2-2024": {"governance": 1.4, "risk": 1.3, "access": 1.4, "operations": 1.2},
        "SDAIA-PDPL": {"privacy": 1.6, "consent": 1.4, "retention": 1.3, "breach": 1.5},
        "ISO-27001-2022": {"governance": 1.3, "risk": 1.4, "improvement": 1.2},
        "DGA-QIYAS-2025": {"strategy": 1.3, "services": 1.3, "data": 1.2},
    }
    weights = weights_by_framework.get(payload.framework_code or "", {})
    denominator = sum(weights.get(key, 1.0) for key in payload.answers)
    numerator = sum(weights.get(key, 1.0) for key, value in payload.answers.items() if value)
    score = round(numerator / denominator * 100) if denominator else 0
    risk = "low" if score >= 80 else "medium" if score >= 55 else "high"
    return {"score": score, "risk": risk, "answered": len(payload.answers), "framework_code": payload.framework_code,
        "method": "framework_weighted" if payload.framework_code in weights_by_framework else "equal_weight",
        "label": "Multazim estimated readiness score"}


@app.post("/v1/evidence", response_model=Evidence, status_code=201)
def create_evidence(payload: EvidenceCreate, user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR))]):
    assert_tenant(payload.organization_id, user)
    evidence = Evidence(id=uuid4(), state="uploaded", **payload.model_dump())
    evidence_store[evidence.id] = evidence
    record_event(user, "evidence.created", "evidence", evidence.id)
    return evidence


@app.post("/v1/evidence/upload", response_model=Evidence, status_code=201)
async def upload_evidence(
    organization_id: Annotated[UUID, Form()], title: Annotated[str, Form(min_length=2, max_length=250)],
    universal_control_id: Annotated[str, Form(min_length=2)], classification: Annotated[str, Form()] = "internal",
    file: UploadFile = File(...),
    user: UserContext = Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR)),
):
    assert_tenant(organization_id, user)
    if classification not in {"public", "internal", "confidential", "restricted"}:
        raise HTTPException(status_code=422, detail="Invalid classification")
    if file.content_type not in ALLOWED_EVIDENCE_TYPES:
        raise HTTPException(status_code=415, detail="Unsupported evidence file type")
    content = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Evidence file exceeds upload limit")
    if not validate_content_type(content, file.content_type):
        raise HTTPException(status_code=415, detail="File content does not match its declared type")
    digest = hashlib.sha256(content).hexdigest()
    evidence_id = uuid4()
    clean, scan_reason = scan_upload(content)
    if not clean:
        raise HTTPException(status_code=422, detail=f"Unsafe evidence file: {scan_reason}")
    safe_suffix = Path(file.filename or "evidence").suffix.lower()[:10]
    put_object(f"{organization_id}/{evidence_id}{safe_suffix}", content)
    evidence = Evidence(id=evidence_id, organization_id=organization_id, title=title,
        universal_control_ids=[universal_control_id], classification=classification, state="under_review",
        sha256=digest, filename=Path(file.filename or "evidence").name, content_type=file.content_type,
        size_bytes=len(content), scan_status="clean")
    evidence_store[evidence.id] = evidence
    record_event(user, "evidence.file_uploaded", "evidence", evidence.id)
    return evidence


@app.get("/v1/evidence", response_model=list[Evidence])
def list_evidence(user: Annotated[UserContext, Depends(user_context)]):
    return [item for item in evidence_store.values() if item.organization_id == user.organization_id]


@app.get("/v1/evidence/{evidence_id}", response_model=Evidence)
def get_evidence(evidence_id: UUID, user: Annotated[UserContext, Depends(user_context)]):
    evidence = evidence_store.get(evidence_id)
    if not evidence:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(evidence.organization_id, user)
    return evidence


@app.patch("/v1/evidence/{evidence_id}", response_model=Evidence)
def update_evidence_lifecycle(evidence_id: UUID, payload: EvidenceLifecycleUpdate,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR))]):
    evidence = evidence_store.get(evidence_id)
    if not evidence:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(evidence.organization_id, user)
    updated = evidence.model_copy(update={"state": payload.state, "expiry_date": payload.expiry_date})
    evidence_store[evidence_id] = updated
    record_event(user, f"evidence.{payload.state}", "evidence", evidence_id)
    capture_snapshot(user, "evidence_updated")
    return updated


@app.post("/v1/evidence/{evidence_id}/analysis", response_model=EvidenceAnalysis)
def analyze_evidence(evidence_id: UUID, payload: EvidenceAnalysisRequest, user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.ASSESSOR))]):
    evidence = evidence_store.get(evidence_id)
    if not evidence:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(evidence.organization_id, user)
    record_event(user, "evidence.analysis_requested", "evidence", evidence_id)
    return EvidenceAnalysis(evidence_id=evidence_id, suggestion="insufficient", confidence=0.35,
        reasoning_ar=f"لم يتم توصيل مزود تحليل المستندات بعد؛ لا يمكن التحقق من المتطلب {payload.requirement_reference} من بيانات الوصف وحدها.",
        citations=[], requires_human_approval=True)


@app.get("/v1/calendar")
def compliance_calendar(user: Annotated[UserContext, Depends(user_context)]):
    return [{"id": str(item.id), "type": "corrective_action", "title": item.title, "date": item.due_date,
        "priority": item.priority, "status": item.status} for item in action_store.values() if item.organization_id == user.organization_id]


@app.get("/v1/audit-log", response_model=list[AuditEvent])
def audit_log(user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.AUDITOR))]):
    return [event for event in audit_events if event.organization_id == user.organization_id]


@app.get("/v1/notifications", response_model=list[Notification])
def notifications(user: Annotated[UserContext, Depends(user_context)], unread_only: bool = False):
    today = date.today()
    for action in action_store.values():
        if action.organization_id == user.organization_id and action.status != "completed" and (action.due_date - today).days <= 14:
            overdue = action.due_date < today
            notify(user, "action_overdue" if overdue else "action_due", "إجراء متأخر" if overdue else "إجراء مستحق قريبًا",
                "Action overdue" if overdue else "Action due soon", "action", action.id, "/gaps",
                "critical" if overdue else "high" if action.priority in {"critical", "high"} else "medium")
    for evidence in evidence_store.values():
        if evidence.organization_id == user.organization_id and evidence.expiry_date and 0 <= (evidence.expiry_date - today).days <= 30:
            notify(user, "evidence_expiring", "دليل يقترب من الانتهاء", "Evidence expiring", "evidence", evidence.id, "/evidence", "high")
    items = [item for item in notification_store.values() if item.organization_id == user.organization_id and item.recipient_id == user.user_id]
    if unread_only:
        items = [item for item in items if item.read_at is None]
    return sorted(items, key=lambda item: item.created_at, reverse=True)


@app.patch("/v1/notifications/{notification_id}/read", response_model=Notification)
def mark_notification_read(notification_id: UUID, user: Annotated[UserContext, Depends(user_context)]):
    item = notification_store.get(notification_id)
    if not item or item.organization_id != user.organization_id or item.recipient_id != user.user_id:
        raise HTTPException(status_code=404, detail="Resource not found")
    updated = item.model_copy(update={"read_at": datetime.now(timezone.utc)})
    notification_store[notification_id] = updated
    return updated


@app.get("/v1/compliance-history", response_model=list[ComplianceSnapshot])
def compliance_history(user: Annotated[UserContext, Depends(user_context)]):
    return sorted([item for item in snapshot_store.values() if item.organization_id == user.organization_id],
        key=lambda item: item.captured_at)


@app.get("/v1/audits/package.json")
def audit_package(user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER, Role.AUDITOR))]):
    payload = {"generated_at": datetime.now(timezone.utc).isoformat(), "organization_id": str(user.organization_id),
        "evidence": [item.model_dump(mode="json") for item in evidence_store.values() if item.organization_id == user.organization_id],
        "actions": [item.model_dump(mode="json") for item in action_store.values() if item.organization_id == user.organization_id],
        "audit_log": [item.model_dump(mode="json") for item in audit_events if item.organization_id == user.organization_id],
        "notice": "Decision-support export; human review and source-document verification remain required."}
    return Response(content=json.dumps(payload, ensure_ascii=False, indent=2), media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=multazim-audit-package.json"})


@app.post("/v1/policies/draft")
def policy_draft(payload: PolicyDraftRequest, user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))]):
    assert_tenant(payload.organization_id, user)
    policy_id = uuid4()
    document = PolicyDocument(id=policy_id, organization_id=payload.organization_id,
        title_ar="مسودة سياسة للاعتماد", title_en="Policy draft for approval", document_type="policy",
        owner=user.user_id, version="0.1", ai_assisted=True, created_at=datetime.now(timezone.utc), updated_at=datetime.now(timezone.utc))
    policy_store[policy_id] = document
    record_event(user, "policy.draft_generated", "policy", policy_id)
    return {"id": policy_id, "policy_type": payload.policy_type, "status": "draft_requires_approval",
        "title_ar": "مسودة سياسة للاعتماد", "notice_ar": "مسودة إرشادية تتطلب مراجعة واعتماد الجهة والمستشار القانوني والأمني حسب الاختصاص.",
        "notice_en": "AI-assisted draft — requires human review and approval",
        "sections": ["الغرض والنطاق", "الأدوار والمسؤوليات", "المتطلبات", "المراقبة والمراجعة", "إدارة الاستثناءات"]}


@app.get("/v1/policies", response_model=list[PolicyDocument])
def list_policies(user: Annotated[UserContext, Depends(user_context)]):
    return [item for item in policy_store.values() if item.organization_id == user.organization_id]


@app.post("/v1/policies", response_model=PolicyDocument, status_code=201)
def create_policy(payload: PolicyDocumentCreate,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))]):
    assert_tenant(payload.organization_id, user)
    now = datetime.now(timezone.utc)
    item = PolicyDocument(id=uuid4(), created_at=now, updated_at=now, **payload.model_dump())
    policy_store[item.id] = item
    record_event(user, "policy.created", "policy", item.id)
    return item


@app.patch("/v1/policies/{policy_id}/transition", response_model=PolicyDocument)
def transition_policy(policy_id: UUID, payload: PolicyTransition,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))]):
    item = policy_store.get(policy_id)
    if not item:
        raise HTTPException(status_code=404, detail="Resource not found")
    assert_tenant(item.organization_id, user)
    allowed = {"draft": {"under_review", "archived"}, "under_review": {"draft", "approved"},
        "approved": {"published", "draft"}, "published": {"superseded", "archived"},
        "superseded": {"archived"}, "archived": set()}
    if payload.status not in allowed[item.status]:
        raise HTTPException(status_code=409, detail=f"Invalid policy transition: {item.status} -> {payload.status}")
    if item.ai_assisted and payload.status in {"approved", "published"} and not item.approver:
        raise HTTPException(status_code=409, detail="AI-assisted policy requires a named human approver")
    updated = item.model_copy(update={"status": payload.status, "updated_at": datetime.now(timezone.utc)})
    policy_store[policy_id] = updated
    record_event(user, f"policy.{payload.status}", "policy", policy_id)
    return updated


@app.post("/v1/knowledge-sources", response_model=KnowledgeSource, status_code=201)
def ingest_knowledge_source(payload: KnowledgeSourceCreate,
    user: Annotated[UserContext, Depends(require_roles(Role.ORG_ADMIN, Role.COMPLIANCE_MANAGER))]):
    if payload.organization_id:
        assert_tenant(payload.organization_id, user)
    normalized = " ".join(payload.content.split())
    chunk_size = 700
    chunks = [normalized[index:index + chunk_size] for index in range(0, len(normalized), chunk_size)]
    item = KnowledgeSource(id=uuid4(), checksum=hashlib.sha256(normalized.encode()).hexdigest(),
        chunks=chunks, ingested_at=datetime.now(timezone.utc), **payload.model_dump())
    knowledge_store[item.id] = item
    record_event(user, "knowledge_source.ingested", "knowledge_source", item.id)
    return item


@app.post("/v1/assistant/query")
def grounded_assistant(payload: AssistantQuery, user: Annotated[UserContext, Depends(user_context)]):
    terms = {term.casefold() for term in payload.question.split() if len(term) > 2}
    candidates = []
    for source in knowledge_store.values():
        if source.organization_id not in {None, user.organization_id} or source.source_status == "demo_unverified":
            continue
        if payload.framework_code and source.framework_code != payload.framework_code:
            continue
        for ordinal, chunk in enumerate(source.chunks):
            score = sum(term in chunk.casefold() for term in terms)
            if score:
                candidates.append((score, source, ordinal, chunk))
    candidates.sort(key=lambda row: row[0], reverse=True)
    citations = [{"source_id": str(source.id), "title": source.title, "url": str(source.source_url),
        "framework": source.framework_code, "version": source.framework_version, "chunk": ordinal,
        "passage": chunk} for _, source, ordinal, chunk in candidates[:3]]
    return {"mode": "deterministic_retrieval", "answer": "Relevant approved source passages are listed in citations. Human interpretation is required."
        if citations else "No approved source passage matched the question; no compliance requirement was inferred.",
        "citations": citations, "uncertainty": "This response is retrieval assistance, not a legal or regulatory conclusion.",
        "prohibited_automatic_actions": ["mark_compliant", "approve_evidence", "approve_mapping", "close_gap", "approve_policy"]}


@app.get("/v1/reports/executive.csv")
def executive_report(user: Annotated[UserContext, Depends(user_context)]):
    summary = dashboard(user)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Multazim Executive Compliance Report", "DEMO / بيانات تجريبية"])
    writer.writerow(["Organization ID", str(user.organization_id)])
    writer.writerow(["Generated At", datetime.now(timezone.utc).isoformat()])
    writer.writerow(["Overall Multazim Estimated Score", summary.overall_score])
    writer.writerow(["Evidence Readiness", summary.evidence_readiness])
    writer.writerow([])
    writer.writerow(["Framework", "Version", "Estimated Score"])
    for framework in summary.framework_scores:
        writer.writerow([framework.name_en, framework.version, framework.score])
    writer.writerow([])
    writer.writerow(["Limitation", "Not an official regulator score or certification"])
    return Response(content="\ufeff" + output.getvalue(), media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=multazim-executive-report.csv"})


@app.get("/v1/reports/executive.xlsx")
def executive_report_xlsx(user: Annotated[UserContext, Depends(user_context)]):
    summary = dashboard(user)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Executive Summary"
    rows = [
        ["Multazim Executive Compliance Report", "DEMO / Decision Support"],
        ["Organization ID", str(user.organization_id)],
        ["Generated At", datetime.now(timezone.utc).isoformat()],
        ["Overall Estimated Score", summary.overall_score],
        ["Evidence Readiness", summary.evidence_readiness],
        [],
        ["Framework", "Version", "Estimated Score"],
        *[[item.name_en, item.version, item.score] for item in summary.framework_scores],
        [],
        ["Limitation", "Not an official regulator score or certification"],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = PatternFill("solid", fgColor="047857")
    for cell in sheet[7]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 48
    sheet.column_dimensions["C"].width = 20
    output = io.BytesIO()
    workbook.save(output)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=multazim-executive-report.xlsx"})


@app.get("/v1/reports/executive.pdf")
def executive_report_pdf(user: Annotated[UserContext, Depends(user_context)], locale: Literal["ar", "en"] = "en"):
    summary = dashboard(user)
    output = io.BytesIO()
    pdf = Canvas(output, pagesize=A4)
    width, height = A4
    pdf.setFillColor(HexColor("#064e3b"))
    pdf.rect(0, height - 110, width, 110, fill=1, stroke=0)
    arabic = locale == "ar"
    body_font = arabic_pdf_font()
    pdf.setFillColor(HexColor("#ffffff"))
    pdf.setFont(body_font, 19)
    title = rtl("تقرير ملتزم التنفيذي للامتثال") if arabic else "Multazim Executive Compliance Report"
    (pdf.drawRightString if arabic else pdf.drawString)(width - 42 if arabic else 42, height - 55, title)
    pdf.setFont(body_font, 8.5)
    disclaimer = rtl("تقرير لدعم القرار - لا يمثل درجة أو شهادة صادرة من جهة تنظيمية") if arabic else "Decision-support report - not a regulator score or certification"
    (pdf.drawRightString if arabic else pdf.drawString)(width - 42 if arabic else 42, height - 78, disclaimer)
    pdf.setFillColor(HexColor("#111827"))
    pdf.setFont(body_font, 12)
    score_label = rtl(f"الجاهزية التقديرية: %{summary.overall_score}") if arabic else f"Overall estimated readiness: {summary.overall_score}%"
    evidence_label = rtl(f"تغطية الأدلة: %{summary.evidence_readiness}") if arabic else f"Evidence coverage: {summary.evidence_readiness}%"
    (pdf.drawRightString if arabic else pdf.drawString)(width - 42 if arabic else 42, height - 150, score_label)
    (pdf.drawRightString if arabic else pdf.drawString)(width - 300 if arabic else 300, height - 150, evidence_label)
    y = height - 195
    pdf.setFont(body_font, 10)
    pdf.drawString(42, y, rtl("الإطار") if arabic else "Framework")
    pdf.drawString(350, y, rtl("الإصدار") if arabic else "Version")
    pdf.drawString(450, y, rtl("الدرجة") if arabic else "Score")
    pdf.setFont(body_font, 9)
    for framework in summary.framework_scores:
        y -= 26
        pdf.drawString(42, y, rtl(framework.name_ar) if arabic else framework.name_en)
        pdf.drawString(350, y, framework.version)
        pdf.drawString(450, y, f"{framework.score}%")
    y -= 42
    pdf.setFont(body_font, 10)
    pdf.drawString(42, y, rtl("إجراءات المعالجة ذات الأولوية") if arabic else "Priority corrective actions")
    pdf.setFont(body_font, 8.5)
    for index, action in enumerate(summary.actions[:5]):
        y -= 22
        action_title = rtl(action.title[:62]) if arabic else f"Corrective action {index + 1}"
        pdf.drawString(42, y, action_title)
        pdf.drawRightString(width - 42, y, f"{action.priority} | {action.due_date.isoformat()}")
    pdf.setFont(body_font, 7.5)
    pdf.setFillColor(HexColor("#6b7280"))
    footer = f"{date.today().isoformat()} | {str(user.organization_id)[:8]} | weighted-status-v1"
    pdf.drawString(42, 38, footer)
    pdf.save()
    return Response(content=output.getvalue(), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=multazim-executive-report-{locale}.pdf"})


@app.get("/v1/frameworks/catalog")
def framework_catalog(
    q: Annotated[str | None, Query(max_length=120)] = None,
    regulator: Annotated[str | None, Query(max_length=40)] = None,
    verification_status: Annotated[str | None, Query(max_length=40)] = None,
):
    root = Path(__file__).resolve().parents[3] / "regulatory_catalog"
    records = []
    for path in sorted(root.rglob("*.json")):
        if path.name == "schema.json":
            continue
        item = json.loads(path.read_text(encoding="utf-8"))
        record = {
            "code": item["code"], "regulator": item["regulator"],
            "version": item["version"], "name_ar": item["name_ar"], "name_en": item["name_en"],
            "official_source": item["official_source"], "verification_status": item["status"],
            "facts": item.get("facts", {}), "controls_count": len(item.get("controls", [])),
        }
        searchable = " ".join(str(record[key]) for key in ("code", "regulator", "name_ar", "name_en", "version")).casefold()
        if q and q.casefold() not in searchable:
            continue
        if regulator and record["regulator"].casefold() != regulator.casefold():
            continue
        if verification_status and record["verification_status"].casefold() != verification_status.casefold():
            continue
        records.append(record)
    return {"count": len(records), "records": records,
        "notice": "Source metadata only; detailed control text requires licensed/official source verification."}


def load_regulatory_journeys() -> list[dict]:
    root = Path(__file__).resolve().parents[3] / "regulatory_journeys"
    return [json.loads(path.read_text(encoding="utf-8")) for path in sorted(root.rglob("*.json")) if path.name != "schema.json"]


@app.get("/v1/journeys")
def regulatory_journeys(q: Annotated[str | None, Query(max_length=120)] = None):
    records = []
    for item in load_regulatory_journeys():
        summary = {
            "code": item["code"], "business_activity": item["business_activity"], "license": item["license"],
            "authority": item["authority"], "platform": item["platform"],
            "requirements_count": len(item["requirements"]),
            "confirmed_count": sum(requirement["status"] == "CONFIRMED_REQUIREMENT" for requirement in item["requirements"]),
        }
        searchable = " ".join((item["code"], item["business_activity"]["name_ar"], item["business_activity"]["name_en"], item["authority"]["code"])).casefold()
        if not q or q.casefold() in searchable:
            records.append(summary)
    return {"count": len(records), "records": records}


@app.get("/v1/journeys/{journey_code}")
def regulatory_journey(journey_code: str):
    journey = next((item for item in load_regulatory_journeys() if item["code"].casefold() == journey_code.casefold()), None)
    if not journey:
        raise HTTPException(status_code=404, detail="Regulatory journey not found")
    return journey


@app.post("/v1/journeys/{journey_code}/readiness")
def journey_readiness(journey_code: str, payload: JourneyReadinessRequest):
    journey = regulatory_journey(journey_code)
    valid_codes = {requirement["code"] for requirement in journey["requirements"]}
    completed = set(payload.completed_requirement_codes)
    unknown = sorted(completed - valid_codes)
    if unknown:
        raise HTTPException(status_code=422, detail={"unknown_requirement_codes": unknown})
    score = sum(requirement["weight"] for requirement in journey["requirements"] if requirement["code"] in completed)
    blockers = [
        {"code": requirement["code"], "title_ar": requirement["title_ar"], "title_en": requirement["title_en"], "verification_status": requirement["status"]}
        for requirement in journey["requirements"] if requirement["code"] not in completed
    ]
    return {
        "journey_code": journey["code"], "score": score,
        "status": "ready_for_submission" if score == 100 else "in_progress" if score else "not_started",
        "completed_count": len(completed), "total_count": len(valid_codes), "blockers": blockers,
        "notice": "Readiness is decision support and does not guarantee regulator approval.",
    }


@app.post("/v1/audits/website")
async def audit_website(payload: WebsiteAuditRequest):
    target = str(payload.url)
    host = payload.url.host
    try:
        addresses = {item[4][0] for item in socket.getaddrinfo(host, payload.url.port or 443)}
    except socket.gaierror as exc:
        raise HTTPException(status_code=422, detail="Website hostname could not be resolved") from exc
    if any(ipaddress.ip_address(address).is_private or ipaddress.ip_address(address).is_loopback
        or ipaddress.ip_address(address).is_link_local or ipaddress.ip_address(address).is_reserved for address in addresses):
        raise HTTPException(status_code=422, detail="Private or reserved network targets are not allowed")
    try:
        async with httpx.AsyncClient(timeout=8, follow_redirects=True, max_redirects=5,
            headers={"User-Agent": "Multazim-Technical-Indicator/1.0"}) as client:
            response = await client.get(target)
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=422, detail="Website could not be safely retrieved") from exc
    content = response.text[:500_000] if "text/html" in response.headers.get("content-type", "") else ""
    lowered = content.casefold()
    links = re.findall(r'''href=["']([^"'#]+)''', content, re.IGNORECASE)[:50]
    indicators = [
        {"code": "https", "passed": response.url.scheme == "https", "detail": str(response.url)},
        {"code": "hsts", "passed": bool(response.headers.get("strict-transport-security")), "detail": "Strict-Transport-Security header"},
        {"code": "csp", "passed": bool(response.headers.get("content-security-policy")), "detail": "Content-Security-Policy header"},
        {"code": "nosniff", "passed": response.headers.get("x-content-type-options", "").casefold() == "nosniff", "detail": "X-Content-Type-Options header"},
        {"code": "privacy_page_signal", "passed": any(term in lowered for term in ("privacy", "الخصوصية", "حماية البيانات")), "detail": "Public privacy-language presence only"},
        {"code": "cookie_notice_signal", "passed": any(term in lowered for term in ("cookie", "cookies", "ملفات تعريف الارتباط")), "detail": "Cookie-language presence only"},
        {"code": "html_language", "passed": bool(re.search(r"<html[^>]+lang=", content, re.IGNORECASE)), "detail": "HTML lang attribute"},
        {"code": "page_title", "passed": bool(re.search(r"<title>.+?</title>", content, re.IGNORECASE | re.DOTALL)), "detail": "Non-empty page title"},
    ]
    return {"url": target, "final_url": str(response.url), "http_status": response.status_code,
        "classification": "TECHNICAL_INDICATORS_ONLY", "indicators": indicators,
        "sampled_links": links[:20], "notice": "These automated technical indicators are not a regulatory compliance conclusion and do not establish PDPL compliance."}
